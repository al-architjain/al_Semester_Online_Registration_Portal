
# Modules used
from django.shortcuts import render
from django.http import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib.auth.models import User, Group
# imports form sorp_app
from . import forms
from . import models
#for excel files.
import openpyxl
#for PasswordChange
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash

#-----------------------------------------------------------#





# Functions

# function to find the group of user
def get_user_group(user):
    g_name = user.groups.values_list('name', flat=True)
    return g_name[0]

#------------------------------------------------------------#





# VIEWS

# for domain redirect
def domain_redirect(request):
    return redirect('login/')





# User Login Authetication
def user_login(request):
    form = forms.login_form()
    logout(request)

    if request.method == "POST":
        form = forms.login_form(request.POST)
        username = password = ''

        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(request, username=username, password=password)

            if (user is not None) and (user.is_active) :
                login(request, user)
                return redirect('/profile/')
            else:
                return render(request, 'sorp_app/login.html', {'form': form, 'invalid':True})

    return render(request, 'sorp_app/login.html', {'form': form, 'invalid': False})





# user_profile
@login_required
def user_profile(request):
    grp = get_user_group(request.user)
    if grp == 'Student':
        user = request.user
        sobj = user.studentinfo
        dobj = models.DocumentInfo.objects.filter(student=sobj, submitted=False)
        subobj = models.Subjects.objects.filter(classname=sobj.ug_class, branch=sobj.ug_branch, semester=sobj.ug_sem)
        robj = models.Result.objects.filter(roll_no=sobj.roll_no).order_by('semester')
        return render(request, 'sorp_app/stu_profile.html', {'sobj': sobj, 'dobj': dobj, 'subobj': subobj, 'robj':robj})

    elif grp == 'Registration Staff':
        uobj = request.user
        dobj = None
        roll_list = models.StudentInfo.objects.filter(reg_staff=request.user)
        # print(roll_list)
        if request.method == 'POST':
            try:
                dobj = models.StudentInfo.objects.get(roll_no=request.POST['d_roll_no'])
                return render(request, 'sorp_app/reg_profile.html', {'uobj': uobj, 'dobj': dobj, 'exist': True, 'tabb': '3','roll_list':roll_list})
            except models.StudentInfo.DoesNotExist:
                dobj = None
                return render(request, 'sorp_app/reg_profile.html', {'uobj': uobj, 'dobj': dobj, 'exist':False, 'tabb':'3','roll_list':roll_list})

        return render(request, 'sorp_app/reg_profile.html',{'uobj': uobj, 'dobj': dobj, 'exist':True, 'tabb':'1','roll_list':roll_list})


    elif grp == 'Library Staff' or grp == 'Hostel Staff' or grp == 'Administration Staff' or grp == 'Department Staff':
        uobj = request.user
        return render(request, 'sorp_app/staff_profile.html', {'uobj': uobj, 'ugrp': grp, 'msg':None})
   

    elif grp == 'Admin':
        return redirect('/admin/')

    else:
        return HttpResponse("You are not suppossed to login from here! ;)")



# create_student user
@login_required
def create_student(request):
    iform = forms.StudentInfoForm()
    fform = forms.StudentFirstFeeForm()
    dobj = models.Documents.objects.all()
    if request.method == "POST":
        iform = forms.StudentInfoForm(request.POST)
        fform = forms.StudentFirstFeeForm(request.POST)

        if (iform.is_valid() and fform.is_valid()) is False:
            print(iform.errors.as_data())
            print(fform.errors.as_data())
            return render(request, 'sorp_app/reg_addstudent.html',
                          {'iform': iform, 'dobj': dobj, 'fform': fform,'uobj':request.user})
        else:
            # iform.reg_staff=request.user
            # print('reg_staff',iform.reg_staff)
            iformm = iform.save(commit=False)
            fformm = fform.save(commit=False)

            stu_name = iform.cleaned_data['name_eng']
            # create user
            username = iform.cleaned_data['roll_no']
            password = iform.cleaned_data['father_name']
            email = iform.cleaned_data['email']
            user = User.objects.create_user(username=username, password=password, email=email)

            # assigning group
            my_group = Group.objects.get(name='Student')
            my_group.user_set.add(user)

            # assgning OnetoOneField and saving
            iformm.user = user
            iformm.save()
            fformm.student = iformm
            fformm.save()

            # document assignment
            for i in range(1, dobj.count() + 1):
                strr = "doc" + str(i)
                submit = request.POST.get(strr)
                if submit is not None:
                    dinfo = models.DocumentInfo(student=iformm, document=dobj[i - 1], submitted=submit)
                    dinfo.save()

            # saving multiple field.
            iform.save_m2m()

            return render(request, 'sorp_app/reg_success.html', {'username':username, 'password':password, 'name':stu_name})

    else:
        return render(request, 'sorp_app/reg_addstudent.html',
                      {'iform': iform, 'dobj': dobj, 'fform': fform,'uobj':request.user})





# Deactivating a student
@login_required
def deactivate(request):
    if request.method == 'GET':
        return redirect('/profile/')
    if request.method == 'POST':
        #to free the allotted roll_no in student_info table
        roll_no = request.POST.get('droll_noo', None)
        stu_info_obj = models.StudentInfo.objects.get(roll_no=roll_no)
        new_roll_no = roll_no if roll_no[-1]=='D' else roll_no+'D'
        stu_info_obj.roll_no = new_roll_no
        stu_info_obj.active_status = False
        stu_info_obj.save()

        # so the person cannot login() and allotted username(roll_no) is freed
        user = User.objects.get(username=roll_no)
        user.is_active = False
        user.username  = new_roll_no
        user.save()

        return HttpResponse("STUDENT DEACTIVATED")





# Successfull Registration Page
def reg_success(request):
    return render(request, 'sorp_app/reg_success.html')





# upload_due
def upload_due(request):
    if request.method =="GET":
        return redirect('/profile/')
    if request.method =="POST":
        grp = get_user_group(request.user)
        path=request.FILES['file']
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        rows = sheet_obj.max_row
        column = sheet_obj.max_column
        i = 2
        while i <= rows:
            roll_obj = sheet_obj.cell(row=i, column=1).value
            fee_obj = sheet_obj.cell(row=i, column=2).value
            try:
                obj1 = models.StudentInfo.objects.get(roll_no=roll_obj)
                obj  = models.Due.objects.get(roll_no=obj1)
            # except models.StudentInfo.DoesNOTExist:
            #     return redirect('/profile/')
                # HttpResponse('some roll no. in file  does not exist')
            except models.StudentInfo.DoesNotExist:
                msg='Error in excel File \n '+str(roll_obj)+' does not exist '
                return redirect('/profile/',{'uobj':request.user,'ugrp':get_user_group(request.user),'msg':msg})
            except models.Due.DoesNotExist:
                if grp == 'Library Staff':
                    obj2 = models.Due(roll_no=obj1, library_due=fee_obj)
                elif grp == 'Administration Staff':
                    obj2 = models.Due(roll_no=obj1, academic_due=fee_obj)
                else:
                    obj2 = models.Due(roll_no=obj1, hostel_due=fee_obj)
                obj2.save()
            else:
                if grp == 'Library Staff':
                    obj.library_due=fee_obj
                elif grp == 'Administration Staff':
                    obj.academic_due=fee_obj
                else:
                    obj.hostel_due=fee_obj
                obj.save()
            i=i+1
    return redirect('/profile/')

@login_required
def upload_sub(request):
    if request.method=="GET":
        return redirect('/profile/')
    if  request.method=="POST":
        path = request.FILES['file']

        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        rows = sheet_obj.max_row
        column = sheet_obj.max_column
        i = 2
        while i <= rows:
            sem=sheet_obj.cell(row=i,column=1).value
            sub_code=sheet_obj.cell(row=i,column=2).value
            sub_name=sheet_obj.cell(row=i,column=3).value
            branch=sheet_obj.cell(row=i,column=4).value
            class_ = sheet_obj.cell(row=i,column=5).value
            sub_L = sheet_obj.cell(row=i, column=7).value
            sub_T = sheet_obj.cell(row=i, column=9).value
            sub_P = sheet_obj.cell(row=i, column=8).value
            sub_C=sheet_obj.cell(row=i,column=6).value
            year_onwards=sheet_obj.cell(row=i,column=10).value
                # obj1=models.Subjects.objects.filter(semester=sem,branch=branch,sub_code=sub_code)
            try:
                obj_class=models.UGClass.objects.get(name=class_)
                obj_branch=models.UGBranch.objects.get(name=branch)
                obj_sub = models.Subjects.objects.get(semester=sem, classname=class_, sub_code=sub_code)
            except models.UGClass.DoesNotExist:
                msg='error in excel file'+str(class_)
                messages.error(request,messages)
                return render(request,'sorp_app/staff_profile.html',{'uobj':request.user, 'ugrp': get_user_group(request.user) ,'msg': msg})
            except  models.UGBranch.DoesNotExist:
                print('error in excel file', branch)
                return redirect('/profile/')
            except models.Subjects.DoesNotExist:
                obj=models.Subjects(semester=sem,branch=obj_branch,sub_code=sub_code,sub_name=sub_name,classname=obj_class,
                                     sub_C=sub_C,sub_L=sub_L,sub_P=sub_P,sub_T=sub_T,year_onwards=year_onwards)
                obj.save()
            else:
                obj = models.Subjects.objects.filter(semester=sem, classname=class_, sub_code=sub_code)
                obj=obj.update(semester=sem,branch=obj_branch,sub_code=sub_code,sub_name=sub_name,classname=obj_class,
                                     sub_C=sub_C,sub_L=sub_L,sub_P=sub_P,sub_T=sub_T,year_onwards=year_onwards)
                # obj.save()
            i=i+1
        return redirect('/profile/')

@login_required
def upload_result(request):
    if request.method=='GET':
        return redirect('/profile/')
    if request.method=='POST':
        path = request.FILES['file']
        obj1=openpyxl.load_workbook(path)
        sheet_obj=obj1.active
        m_row=sheet_obj.max_row

        i=2
        while(i<=m_row):
            sem=sheet_obj.cell(row=i,column=1).value
            roll_no=sheet_obj.cell(row=i,column=2).value
            sgpi=sheet_obj.cell(row=i,column=3).value
            cgpi=sheet_obj.cell(row=i,column=3).value
            # active_backlogs=sheet_obj.cell(row=i,column=4).value

            obj=models.StudentInfo.objects.filter(roll_no=roll_no)
            print('roll_no ',obj)
            obj1=models.Result.objects.filter(roll_no=obj)
            if(not obj):
                print('error type1')
                return redirect('/profile/')
            # elif(not obj1):
            #     print('error check2')
            #     obj2=Result(roll_no=obj,semester=sem,sgpi=sgpi,cgpi=cgpi)
            #     obj2.save()
            else:
                print('error check 3')
                obj1.update(semester=sem,roll_no=obj,sgpi=sgpi,cgpi=cgpi)
            i=i+1
        return redirect('/profile/')



@login_required()
def change_password(request):
    if request.method == 'POST':
        pcform = PasswordChangeForm(request.user, request.POST)
        if pcform.is_valid():
            user = pcform.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('sorp_app:user_profile_page')

        else:
            messages.error(request, 'Please correct the error below.')
            return render(request, 'sorp_app/stu_password_change.html', {'pcform': pcform})

    else:
        pcform = PasswordChangeForm(request.user)
        return render(request, 'sorp_app/stu_password_change.html', {'pcform': pcform})







@login_required
def update_student(request):
    if request.method == 'GET':
        return redirect('/profile/')

    if request.method == 'POST':
        roll_no = request.POST.get('rollno_to_update', None)
        stu_obj = models.StudentInfo.objects.get(roll_no=roll_no)
        fee_obj = models.StudentFirstFeeStatus.objects.get(student=stu_obj)
        si_data={
            'institute':stu_obj.institute,
            'name_eng':stu_obj.name_eng,
            'name_hindi':stu_obj.name_hindi,
            'email':stu_obj.email,
            'gender':stu_obj.gender,
            'dob':stu_obj.dob,
            'religion':stu_obj.religion,
            'category_main':stu_obj.category_main,
            'contact':stu_obj.contact,
            'aadhar_no':stu_obj.aadhar_no,
            'area':stu_obj.area,
            'b_country':stu_obj.b_country,
            'b_state':stu_obj.b_state,
            'nearest_rs':stu_obj.nearest_rs,
            'corr_addr':stu_obj.corr_addr,
            'perm_addr':stu_obj.perm_addr,
            'jee_roll_no':stu_obj.jee_roll_no,
            'jee_score':stu_obj.jee_score,
            'jee_ai_rank':stu_obj.jee_ai_rank,
            'jee_cat_rank':stu_obj.jee_cat_rank,
            'category_admission':stu_obj.category_admission,
            'int_country':stu_obj.int_country,
            'int_state':stu_obj.int_state,
            'int_percentage':stu_obj.int_percentage,
            'int_pass_year':stu_obj.int_pass_year,
            'int_school_type':stu_obj.int_school_type,
            'int_school_area':stu_obj.int_school_area,
            'int_school_name':stu_obj.int_school_name,
            'int_school_board':stu_obj.int_school_board,
            'ug_class':stu_obj.ug_class,
            'ug_branch':stu_obj.ug_branch,
            'hosteler':stu_obj.hosteler,
            'hostel_name':stu_obj.hostel_name,
            'entry_no':stu_obj.entry_no,
            'roll_no':stu_obj.roll_no,
            'reg_no':stu_obj.reg_no,
            'father_name':stu_obj.father_name,
            'father_contact':stu_obj.father_contact,
            'father_email':stu_obj.father_email,
            'mother_name':stu_obj.mother_name,
            'mother_contact':stu_obj.mother_contact,
            'mother_email':stu_obj.mother_email,
            'guardian_name':stu_obj.guardian_name,
            'guardian_contact':stu_obj.guardian_contact,
            'guardian_email':stu_obj.guardian_email,
            'fee_waiver':stu_obj.fee_waiver,
            'family_income':stu_obj.family_income,

        }

        sfi_data = {
            'fee_josaa_amount': fee_obj.fee_josaa_amount,
            'fee_josaa_date': fee_obj.fee_josaa_date,
            'fee_NITH_amount': fee_obj.fee_NITH_amount,
            'fee_nith_date': fee_obj.fee_nith_date,
            'fee_nith_receipt_no': fee_obj.fee_nith_receipt_no,
        }

        iform = forms.StudentInfoForm(initial=si_data)
        fform = forms.StudentFirstFeeForm(initial=sfi_data)
        dobj = models.Documents.objects.all()
        print(stu_obj.id)
        return render(request, 'sorp_app/reg_updatestudent.html',
                          {'iform': iform, 'dobj': dobj, 'fform': fform, 'uobj': request.user, 'stu_id':stu_obj.id})

    else:
        return HttpResponse('ERRoR!')





@login_required
def update_student_info(request):
    if request.method == 'GET':
        return redirect('/profile/')
    if request.method == 'POST':
        # iform = forms.StudentInfoForm(request.POST)
        stu_obj = models.StudentInfo.objects.get(id=request.POST.get('stu_id',None))
        user = stu_obj.user

        stu_obj.roll_no=request.POST.get('roll_no',None)
        stu_obj.institute=request.POST.get('institute',None)
        stu_obj.name_eng=request.POST.get('name_eng',None)
        stu_obj.name_hindi=request.POST.get('name_hindi',None)
        stu_obj.email=request.POST.get('email',None)
        stu_obj.gender=request.POST.get('gender',None)
        stu_obj.dob=request.POST.get('dob',None)
        stu_obj.religion=request.POST.get('religion',None)
        stu_obj.category_main=models.Category.objects.get(name=request.POST.get('category_main',None))
        stu_obj.contact=request.POST.get('contact',None)
        stu_obj.aadhar_no=request.POST.get('aadhar_no',None)
        stu_obj.area=request.POST.get('area',None)
        stu_obj.b_country=request.POST.get('b_country',None)
        stu_obj.b_state=request.POST.get('b_state',None)
        stu_obj.nearest_rs=request.POST.get('nearest_rs',None)
        stu_obj.corr_addr=request.POST.get('corr_addr',None)
        stu_obj.perm_addr=request.POST.get('perm_addr',None)
        stu_obj.jee_roll_no=request.POST.get('jee_roll_no',None)
        stu_obj.jee_score=request.POST.get('jee_score',None)
        stu_obj.jee_ai_rank=request.POST.get('jee_ai_rank',None)
        stu_obj.jee_cat_rank=request.POST.get('jee_cat_rank',None)
        stu_obj.category_admission=models.Category.objects.get(name=request.POST.get('category_admission',None))
        stu_obj.int_country=request.POST.get('int_country',None)
        stu_obj.int_state=request.POST.get('int_state',None)
        stu_obj.int_percentage=request.POST.get('int_percentage',None)
        stu_obj.int_pass_year=request.POST.get('int_pass_year',None)
        stu_obj.int_school_type=request.POST.get('int_school_type',None)
        stu_obj.int_school_area=request.POST.get('int_school_area',None)
        stu_obj.int_school_name=request.POST.get('int_school_name',None)
        stu_obj.int_school_board=models.Board.objects.get(name=request.POST.get('int_school_board',None))
        stu_obj.ug_class=models.UGClass.objects.get(name=request.POST.get('ug_class',None))
        stu_obj.ug_branch=models.UGBranch.objects.get(name=request.POST.get('ug_branch',None))
        # stu_obj.hosteler=request.POST.get('hosteler',None)
        stu_obj.hostel_name=request.POST.get('hostel_name',None)
        stu_obj.entry_no=request.POST.get('entry_no',None)
        stu_obj.reg_no=request.POST.get('reg_no',None)
        stu_obj.father_name=request.POST.get('father_name',None)
        stu_obj.father_contact=request.POST.get('father_contact',None)
        stu_obj.father_email=request.POST.get('father_email',None)
        stu_obj.mother_name=request.POST.get('mother_name',None)
        stu_obj.mother_contact=request.POST.get('mother_contact',None)
        stu_obj.mother_email=request.POST.get('mother_email',None)
        stu_obj.guardian_name=request.POST.get('guardian_name',None)
        stu_obj.guardian_contact=request.POST.get('guardian_contact',None)
        stu_obj.guardian_email=request.POST.get('guardian_email',None)
        stu_obj.fee_waiver=request.POST.get('fee_waiver',None)
        stu_obj.family_income=request.POST.get('family_income',None)
        # print(request.POST.get('hosteler',None))
        stu_obj.save()
        user.username = request.POST.get('roll_no', None)
        user.set_password(request.POST.get('father_name',None))
        user.save()

        fee_obj=models.StudentFirstFeeStatus.objects.get(student=stu_obj)
        fee_obj.fee_josaa_amount=request.POST.get('fee_josaa_amount',None)
        fee_obj.fee_josaa_date=request.POST.get('fee_josaa_date',None)
        fee_obj.fee_NITH_amount=request.POST.get('fee_NITH_amount',None)
        fee_obj.fee_nith_date=request.POST.get('fee_nith_date',None)
        fee_obj.fee_nith_receipt_no=request.POST.get('fee_nith_receipt_no',None)
        fee_obj.save()



        # if iform.is_valid():
        #     return HttpResponse('Valid form!')
        # else:
        return render(request,'sorp_app/reg_success.html',{'name':stu_obj.name_eng, 'username': user.username, 'password':stu_obj.father_name})

